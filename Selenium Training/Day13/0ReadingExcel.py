#------------------------Data Driven Testing-------------------------------

# openpyxl - we can work with excel files (.xslx)

# 1) Read data from excel           #sheet.cell(r,c).value
# 2) How to write data into excel   #sheet.cell(r,c).value="Hola"
# 3) Data driven test case


#File--->Workbook--->Sheets--->Rows--->Cells

#---ReadingExcel----

import openpyxl

#File--->Workbook--->Sheets--->Rows--->Cells

file="C:/Users/karth/OneDrive/Documents/Selenium practice/data.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook["Sheet1"]
rows=sheet.max_row #count no of rows in excel sheet
columns=sheet.max_column #count no of columns in excel sheet

# Reading all the rows & columns from excel sheet
for r in range(1,rows+1):
    for c in range(1,columns+1):
        print(sheet.cell(r,c).value,end='  ')
    print()