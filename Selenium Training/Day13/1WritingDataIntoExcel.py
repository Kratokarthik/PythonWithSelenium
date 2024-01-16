import openpyxl

#File--->Workbook--->Sheets--->Rows--->Cells

# writing same data in all fields
# file="C:/Users/karth/OneDrive/Documents/Selenium practice/testdata.xlsx"
# workbook=openpyxl.load_workbook(file)
# #sheet=workbook["Data"]    #sheet name "Data"
# sheet=workbook.active      #get active(only) sheet from excel
# for r in range(1,6):
#     for c in range(1,4):
#         sheet.cell(r,c).value="Hola"
#
# workbook.save(file)

#Multiple data
file="C:/Users/karth/OneDrive/Documents/Selenium practice/testdata1.xlsx"
workbook=openpyxl.load_workbook(file)
#sheet=workbook["Data"]    #sheet name "Data"
sheet=workbook.active      #get active(only) sheet from excel

sheet.cell(1,1).value=123
sheet.cell(1,2).value="John"

sheet.cell(2,1).value=567
sheet.cell(2,2).value="Smith"

sheet.cell(3,1).value=324
sheet.cell(3,2).value="Dannny"

sheet.cell(4,1).value=124
sheet.cell(4,2).value="Drake"

workbook.save(file)   #save the file after entering data